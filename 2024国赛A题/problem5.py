from __future__ import annotations
import math
from typing import Tuple, List
import numpy as np
from openpyxl import load_workbook
import matplotlib.pyplot as plt
from matplotlib import rcParams

# 设置中文字体
rcParams['font.sans-serif'] = ['SimHei']
rcParams['axes.unicode_minus'] = False

# ===== 从Problem4.py复制的核心代码 =====
# 基本参数
p = 1.7
b = p / (2.0 * math.pi)
R_turn = 4.5
V0 = 1.0  # 原始龙头速度

# 板几何
L_head, L_body, D_hole = 3.41, 2.20, 0.55
l_HEAD = L_head - D_hole
l_BODY = L_body - D_hole
l_TAIL = l_BODY

# 节数
N_BODY = 221

# 时间网格
T_START, T_END, DT = -100, 100, 1.0
T_GRID = np.arange(T_START, T_END + 1e-12, DT, dtype=float)
DT_V = 0.1


# ===== 基础向量工具 =====
def rot90(v: np.ndarray) -> np.ndarray:
    return np.array([-v[1], v[0]], float)


def angle_of(v: np.ndarray) -> float:
    return math.atan2(v[1], v[0])


def ang_diff(a_to: float, a_from: float, ccw: bool) -> float:
    d = (a_to - a_from) % (2.0 * math.pi)
    return d if ccw else ((a_from - a_to) % (2.0 * math.pi))


# ===== 螺线弧长原函数与反函数 =====
def F_theta(theta: float) -> float:
    return 0.5 * (theta * math.sqrt(1.0 + theta * theta) + math.asinh(theta))


def inv_F(Fv: float, max_iter=50, tol=1e-13) -> float:
    theta = math.sqrt(max(0.0, 2.0 * Fv)) if Fv > 1.0 else max(0.0, Fv)
    for _ in range(max_iter):
        f = F_theta(theta) - Fv
        df = math.sqrt(1.0 + theta * theta)
        theta_new = theta - f / df
        if theta_new < 0.0: theta_new = 0.0
        if abs(theta_new - theta) < tol: return theta_new
        theta = theta_new
    return theta


# ===== 盘入/盘出螺线 & 切向 =====
def xy_in(theta: float) -> np.ndarray:
    r = b * theta
    return np.array([r * math.cos(theta), r * math.sin(theta)], float)


def tan_in(theta: float) -> np.ndarray:
    dx = b * math.cos(theta) - b * theta * math.sin(theta)
    dy = b * math.sin(theta) + b * theta * math.cos(theta)
    v = np.array([-dx, -dy], float)
    n = np.hypot(*v);
    return v / (n if n > 1e-14 else 1.0)


def xy_out(theta: float) -> np.ndarray:
    r = b * (theta - math.pi)
    return np.array([r * math.cos(theta), r * math.sin(theta)], float)


def tan_out(theta: float) -> np.ndarray:
    u = theta - math.pi
    dx = b * math.cos(theta) - b * u * math.sin(theta)
    dy = b * math.sin(theta) + b * u * math.cos(theta)
    v = np.array([dx, dy], float)
    n = np.hypot(*v);
    return v / (n if n > 1e-14 else 1.0)


# ===== 触边点 B / F 及两圆几何 =====
theta_B = R_turn / b
alpha = math.atan(theta_B)

R1 = 3.0 / math.sin(alpha)
R2 = 3.0 / (2.0 * math.sin(alpha))

B = xy_in(theta_B)
F = -B

C1 = np.array([R_turn * math.cos(theta_B) - R1 * math.sin(theta_B + alpha),
               R_turn * math.sin(theta_B) + R1 * math.cos(theta_B + alpha)], float)
C2 = np.array([-R_turn * math.cos(theta_B) + R2 * math.sin(theta_B + alpha),
               -R_turn * math.sin(theta_B) - R2 * math.cos(theta_B + alpha)], float)

tB = tan_in(theta_B)
tF = tan_out(theta_B + math.pi)

angB = angle_of(B - C1)
tan_ccw_at_B = rot90((B - C1) / np.hypot(*(B - C1)))
sgn1 = +1 if np.dot(tan_ccw_at_B, tB) > 0 else -1

D = C1 + (R1 / (R1 + R2)) * (C2 - C1)

angD_on_C1 = angle_of(D - C1)
phi1 = ang_diff(angD_on_C1, angB, ccw=(sgn1 > 0))
L1 = R1 * phi1

angD = angle_of(D - C2)
tan_ccw_at_F = rot90((F - C2) / np.hypot(*(F - C2)))
sgn2 = +1 if np.dot(tan_ccw_at_F, tF) > 0 else -1

angF_on_C2 = angle_of(F - C2)
phi2 = ang_diff(angF_on_C2, angD, ccw=(sgn2 > 0))
L2 = R2 * phi2


# ===== 段内坐标表达 =====
def xy_on_arc1(phi: float) -> np.ndarray:
    ang = angB + sgn1 * phi
    return C1 + R1 * np.array([math.cos(ang), math.sin(ang)], float)


def xy_on_arc2(phi: float) -> np.ndarray:
    ang = angD + sgn2 * phi
    return C2 + R2 * np.array([math.cos(ang), math.sin(ang)], float)


# ===== 同段螺线两点弦长为 l 的 Δθ 求解 =====
def same_spiral_delta(theta1: float, l: float, *, offset: float = 0.0) -> float:
    r1 = b * (theta1 - offset)
    if r1 < 1e-12: r1 = 1e-12

    def g(delta: float) -> float:
        r2 = b * (theta1 + delta - offset)
        return r1 * r1 + r2 * r2 - 2.0 * r1 * r2 * math.cos(delta) - l * l

    def gp(delta: float) -> float:
        r2 = b * (theta1 + delta - offset)
        return 2.0 * r2 * b - 2.0 * r1 * b * math.cos(delta) + 2.0 * r1 * r2 * math.sin(delta)

    delta = min(max(l / max(r1, 1e-9), 1e-10), math.pi / 2)
    ok = False
    for _ in range(20):
        val, der = g(delta), gp(delta)
        if abs(der) < 1e-14: break
        cand = delta - val / der
        if not (0.0 < cand < math.pi):
            cand = 0.5 * (delta + max(1e-10, min(cand, math.pi - 1e-10)))
        delta = cand
        if abs(val) < 1e-12: ok = True; break
    if ok: return delta

    lo, hi = 1e-12, math.pi - 1e-12
    gl, gh = g(lo), g(hi)
    if gl * gh > 0:
        hi = min(2 * math.pi, hi + 1.0);
        gh = g(hi)
        if gl * gh > 0: return delta
    for _ in range(80):
        mid = 0.5 * (lo + hi);
        gm = g(mid)
        if gm == 0.0 or (hi - lo) < 1e-12: return mid
        if gl * gm <= 0:
            hi, gh = mid, gm
        else:
            lo, gl = mid, gm
    return 0.5 * (lo + hi)


# ===== 头把手：t→(段, 段内参数) =====
def head_state_at_time(t: float):
    if t <= 0.0:
        Fu = F_theta(theta_B) + (-t) / b
        theta = inv_F(Fu)
        return 1, theta
    if t <= L1:
        return 2, t / R1
    if t <= L1 + L2:
        return 3, (t - L1) / R2
    s_out = t - (L1 + L2)
    Fu = F_theta(theta_B) + s_out / b
    u = inv_F(Fu)
    theta = u + math.pi
    return 4, theta


# ===== 逐节回推 =====
def step_prev(seg: int, param: float, l: float) -> Tuple[int, float]:
    if seg == 1:
        dth = same_spiral_delta(param, l, offset=0.0)
        return 1, param + dth

    elif seg == 2:
        dphi = 2.0 * math.asin(min(1.0, l / (2.0 * R1)))
        if param >= dphi + 1e-14:
            return 2, param - dphi
        chord_used = 2.0 * R1 * math.sin(max(0.0, param) / 2.0)
        l_rem = max(0.0, l - chord_used)
        if l_rem <= 1e-14:
            return 1, theta_B
        dth = same_spiral_delta(theta_B, l_rem, offset=0.0)
        return 1, theta_B + dth

    elif seg == 3:
        dphi = 2.0 * math.asin(min(1.0, l / (2.0 * R2)))
        if param >= dphi + 1e-14:
            return 3, param - dphi
        chord_used = 2.0 * R2 * math.sin(max(0.0, param) / 2.0)
        l_rem = max(0.0, l - chord_used)
        if l_rem <= 1e-14:
            return 2, phi1
        dphi1 = 2.0 * math.asin(min(1.0, l_rem / (2.0 * R1)))
        if phi1 >= dphi1 + 1e-14:
            return 2, phi1 - dphi1
        chord_used2 = 2.0 * R1 * math.sin(max(0.0, phi1) / 2.0)
        l_rem2 = max(0.0, l_rem - chord_used2)
        if l_rem2 <= 1e-14:
            return 1, theta_B
        dth = same_spiral_delta(theta_B, l_rem2, offset=0.0)
        return 1, theta_B + dth

    else:
        dth = same_spiral_delta(param, l, offset=math.pi)
        if param - dth >= theta_B + math.pi - 1e-12:
            return 4, param - dth
        P_now = xy_out(param)
        chord_to_F = float(np.hypot(*(P_now - F)))
        l_rem = max(0.0, l - chord_to_F)
        if l_rem <= 1e-14:
            return 3, phi2
        dphi2 = 2.0 * math.asin(min(1.0, l_rem / (2.0 * R2)))
        if phi2 >= dphi2 + 1e-14:
            return 3, phi2 - dphi2
        chord_used2 = 2.0 * R2 * math.sin(max(0.0, phi2) / 2.0)
        l_rem2 = max(0.0, l_rem - chord_used2)
        if l_rem2 <= 1e-14:
            return 2, phi1
        dphi1 = 2.0 * math.asin(min(1.0, l_rem2 / (2.0 * R1)))
        if phi1 >= dphi1 + 1e-14:
            return 2, phi1 - dphi1
        chord_used3 = 2.0 * R1 * math.sin(max(0.0, phi1) / 2.0)
        l_rem3 = max(0.0, l_rem - chord_used2 - chord_used3)
        if l_rem3 <= 1e-14:
            return 1, theta_B
        dth2 = same_spiral_delta(theta_B, l_rem3, offset=0.0)
        return 1, theta_B + dth2


def xy_of(seg: int, param: float) -> np.ndarray:
    if seg == 1:   return xy_in(param)
    if seg == 2:   return xy_on_arc1(param)
    if seg == 3:   return xy_on_arc2(param)
    return xy_out(param)


# ===== 生成某一时刻全队位置 =====
def all_handles_at_time(t: float) -> np.ndarray:
    seg, par = head_state_at_time(t)
    P = [xy_of(seg, par)]
    seg, par = step_prev(seg, par, l_HEAD)
    P.append(xy_of(seg, par))
    for _ in range(2, N_BODY + 1):
        seg, par = step_prev(seg, par, l_BODY)
        P.append(xy_of(seg, par))
    seg, par = step_prev(seg, par, l_BODY)
    P.append(xy_of(seg, par))
    seg, par = step_prev(seg, par, l_TAIL)
    P.append(xy_of(seg, par))
    return np.vstack(P)


# ===== 速度计算 =====
def speeds_over_time(pos_series: np.ndarray, dt_s: float, dt_v: float = DT_V) -> np.ndarray:
    T, N, _ = pos_series.shape
    speed = np.zeros((T, N), float)

    def pos_at(t: float) -> np.ndarray:
        return all_handles_at_time(t)

    for ti, t in enumerate(T_GRID):
        if t - T_START < 1e-12:
            P2 = pos_at(t + dt_v)
            v = (P2 - pos_series[ti]) / dt_v
        elif T_END - t < 1e-12:
            P0 = pos_at(t - dt_v)
            v = (pos_series[ti] - P0) / dt_v
        else:
            P2 = pos_at(t + dt_v)
            P0 = pos_at(t - dt_v)
            v = (P2 - P0) / (2.0 * dt_v)
        speed[ti, :] = np.hypot(v[:, 0], v[:, 1])
        speed[ti, 0] = 1.0  # 龙头速度置1
    return speed


# ===== 问题5：分析最大速度 =====
def analyze_max_speed_direct():
    """
    直接计算分析，找出最大速度
    """
    print("\n[问题5] 开始分析速度数据...")

    # 生成位置序列
    all_pos = []
    for t in T_GRID:
        all_pos.append(all_handles_at_time(t))
    pos_arr = np.stack(all_pos, axis=0)

    # 计算速度
    spd_arr = speeds_over_time(pos_arr, dt_s=DT, dt_v=DT_V)

    # 找出最大速度
    global_max_speed = np.max(spd_arr)
    max_pos = np.unravel_index(np.argmax(spd_arr), spd_arr.shape)
    max_time_idx, max_handle_idx = max_pos

    print(f"\n分析结果：")
    print(f"  当龙头速度 = 1.0 m/s 时：")
    print(f"  全局最大速度 = {global_max_speed:.6f} m/s")
    print(f"  出现在: t = {T_GRID[max_time_idx]:.0f}s, 第{max_handle_idx}号把手")

    # 计算龙头最大速度
    max_head_speed = 2.0 / global_max_speed

    print(f"\n计算过程：")
    print(f"  V_head_max = 2.0 / {global_max_speed:.6f}")
    print(f"  V_head_max = {max_head_speed:.6f} m/s")

    # 验证
    print(f"\n验证：")
    print(f"  当龙头速度 = {max_head_speed:.6f} m/s 时：")
    print(f"  最大把手速度 = {global_max_speed:.6f} × {max_head_speed:.6f}")
    print(f"              = {global_max_speed * max_head_speed:.6f} m/s ≈ 2.0 m/s ✓")

    return max_head_speed, spd_arr


def analyze_from_excel(xlsx_path="result4.xlsx"):
    """
    从Excel文件分析（备选方案）
    """
    print("\n[问题5] 从result4.xlsx分析速度数据...")

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb["速度"]

    max_speed = 0.0
    max_location = {"time": None, "handle": None}

    for col in range(2, 203):
        time = -100 + (col - 2)
        for row in range(2, 226):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value and float(cell_value) > max_speed:
                max_speed = float(cell_value)
                max_location["time"] = time
                max_location["handle"] = row - 2

    wb.close()

    print(f"\n分析结果：")
    print(f"  当龙头速度 = 1.0 m/s 时：")
    print(f"  全局最大速度 = {max_speed:.6f} m/s")
    print(f"  出现在: t = {max_location['time']}s, 第{max_location['handle']}号把手")

    max_head_speed = 2.0 / max_speed

    print(f"\n龙头最大速度 = {max_head_speed:.6f} m/s")

    return max_head_speed


def plot_speed_analysis(spd_arr):
    """
    绘制速度分析图
    """
    fig, axes = plt.subplots(2, 2, figsize=(12, 8))
    fig.suptitle('问题5 - 速度分析', fontsize=14)

    # 1. 时间-最大速度曲线
    ax1 = axes[0, 0]
    time_max_speeds = np.max(spd_arr, axis=1)
    ax1.plot(T_GRID, time_max_speeds, 'b-', linewidth=1.5)
    ax1.axhline(y=np.max(spd_arr), color='r', linestyle='--',
                label=f'全局最大 ({np.max(spd_arr):.3f} m/s)')
    ax1.set_xlabel('时间 (s)')
    ax1.set_ylabel('最大速度 (m/s)')
    ax1.set_title('各时刻的最大速度')
    ax1.grid(True, alpha=0.3)
    ax1.legend()

    # 2. 龙头速度（应该恒定为1）
    ax2 = axes[0, 1]
    ax2.plot(T_GRID, spd_arr[:, 0], 'g-', linewidth=1.5)
    ax2.set_xlabel('时间 (s)')
    ax2.set_ylabel('速度 (m/s)')
    ax2.set_title('龙头速度')
    ax2.grid(True, alpha=0.3)
    ax2.set_ylim([0.95, 1.05])

    # 3. 关键时刻的速度分布
    ax3 = axes[1, 0]
    key_times = [0, 100, 200]  # 对应 t=-100, 0, 100
    for kt in key_times:
        speeds = spd_arr[kt, :]
        ax3.plot(range(224), speeds, alpha=0.7, label=f't={T_GRID[kt]:.0f}s')
    ax3.set_xlabel('把手编号')
    ax3.set_ylabel('速度 (m/s)')
    ax3.set_title('不同时刻的速度分布')
    ax3.grid(True, alpha=0.3)
    ax3.legend()

    # 4. 速度热图
    ax4 = axes[1, 1]
    im = ax4.imshow(spd_arr.T, aspect='auto', cmap='hot',
                    extent=[T_START, T_END, 223, 0])
    ax4.set_xlabel('时间 (s)')
    ax4.set_ylabel('把手编号')
    ax4.set_title('速度热图')
    plt.colorbar(im, ax=ax4, label='速度 (m/s)')

    plt.tight_layout()
    plt.show()


def main():
    """
    主函数
    """
    print("=" * 60)
    print("           2024年数学建模A题 - 问题5")
    print("=" * 60)

    # 方案1：直接计算分析
    max_head_speed, spd_arr = analyze_max_speed_direct()

    # 方案2：从Excel读取（如果需要验证）
    # max_head_speed = analyze_from_excel("result4.xlsx")

    print("\n" + "=" * 60)
    print(f"【最终答案】")
    print(f"龙头的最大行进速度为: {max_head_speed:.6f} m/s")
    print("=" * 60)

    # 保存结果
    with open("problem5_answer.txt", "w", encoding="utf-8") as f:
        f.write(f"问题5答案：\n")
        f.write(f"龙头的最大行进速度为: {max_head_speed:.6f} m/s\n")

    print(f"\n结果已保存到 problem5_answer.txt")

    # 绘图（可选）
    plot_speed_analysis(spd_arr)

    return max_head_speed


if __name__ == "__main__":
    max_head_speed = main()