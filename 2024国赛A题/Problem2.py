import math
from typing import Tuple, List
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon
plt.rcParams['font.family'] = 'SimHei'
plt.rcParams['axes.unicode_minus'] = False

# --------------- 基本参数 ---------------
PITCH = 0.55                        # 螺距 p（m）
B = PITCH / (2.0 * math.pi)         # b = p/(2π)
THETA0 = 2.0 * math.pi * 16         # t=0 时龙头角（第 16 圈 A 点）
V = 1.0                             # 龙头弧长速度（m/s）

# 板长（米）与“同板前后把手中心距离” l = L - 0.55
L_HEAD, L_BODY, L_TAIL = 3.41, 2.20, 2.20
D_HOLE = 0.55
l_HEAD = L_HEAD - D_HOLE            # 2.86
l_BODY = L_BODY - D_HOLE            # 1.65
l_TAIL = L_TAIL - D_HOLE            # 1.65

# 时间与输出
T0, T1, DT = 0.0, 300.0, 1.0

# --------------- 弧长原函数与反解 ---------------
def F_theta(theta: float) -> float:
    """F(θ) = 0.5*( θ*sqrt(1+θ^2) + asinh(θ) )"""
    return 0.5 * (theta * math.sqrt(1.0 + theta * theta) + math.asinh(theta))

def inv_F(Fv: float, max_iter: int = 40, tol: float = 1e-13) -> float:
    """牛顿法反解 θ，使 F(θ) = Fv（θ≥0，F 单调增）"""
    theta = math.sqrt(2.0 * Fv) if Fv > 1.0 else Fv
    for _ in range(max_iter):
        f = F_theta(theta) - Fv
        df = math.sqrt(1.0 + theta * theta)
        theta_new = theta - f / df
        if theta_new < 0.0:
            theta_new = 0.0
        if abs(theta_new - theta) < tol:
            theta = theta_new
            break
        theta = theta_new
    return theta

def head_polar_at_time(t: float) -> Tuple[float, float]:
    """给定 t，求龙头极坐标 (r, θ)。顺时针盘入：F(θ)=F(θ0)-t/b"""
    F0 = F_theta(THETA0)
    theta = inv_F(F0 - t / B)      # v=1
    r = B * theta
    return r, theta


# --------------- 求同板 Δ（最近物理解，Δ∈(0, π)） ---------------
def solve_delta_same_board(theta1: float, l: float, b: float = B) -> float:
    """
    已知同一条螺线上的“前把手”角 θ1（r1 = bθ1），求同板“后把手”使两点直线距离为 l。
    设 Δ = θ2 - θ1 > 0，r2 = b(θ1 + Δ)，方程：
      g(Δ) = r1^2 + r2^2 - 2 r1 r2 cosΔ - l^2 = 0
    取最近的物理解 Δ ∈ (0, π)。先牛顿（带投影/阻尼），失败退二分。
    """
    r1 = b * theta1

    def g(delta: float) -> float:
        r2 = b * (theta1 + delta)
        return r1*r1 + r2*r2 - 2.0*r1*r2*math.cos(delta) - l*l

    def gprime(delta: float) -> float:
        # r2 = b(θ1+Δ)
        r2 = b * (theta1 + delta)
        # dg/dΔ = 2 r2 b - 2 r1 b cosΔ + 2 r1 r2 sinΔ
        return 2.0*r2*b - 2.0*r1*b*math.cos(delta) + 2.0*r1*r2*math.sin(delta)

    # 牛顿初值：小角近似 Δ ≈ l / max(r1, 1e-9)
    delta = min(max(l / max(r1, 1e-9), 1e-10), math.pi/2)
    ok = False
    for _ in range(20):
        val = g(delta)
        der = gprime(delta)
        if abs(der) < 1e-14:
            break
        # 阻尼牛顿 + 投影到 (0, π)
        cand = delta - val/der
        if not (0.0 < cand < math.pi):
            cand = 0.5*(delta + max(1e-10, min(cand, math.pi-1e-10)))
        delta = cand
        if abs(val) < 1e-12:
            ok = True
            break
    if ok:
        return delta

    # 二分：在 (0, π) 内找根
    low, high = 1e-12, math.pi - 1e-12
    gl, gh = g(low), g(high)   # g(0+)≈-l^2<0, g(π-)通常>0
    if gl*gh > 0:
        # 极端场景：略扩上界再试
        high = min(2.0*math.pi, high + 1.0)
        gh = g(high)
        if gl*gh > 0:
            return delta  # 退回牛顿最后值
    for _ in range(80):
        mid = 0.5*(low + high)
        gm = g(mid)
        if gm == 0.0 or (high - low) < 1e-12:
            return mid
        if gl*gm <= 0.0:
            high, gh = mid, gm
        else:
            low, gl = mid, gm
    return 0.5*(low + high)


def step_same_board(theta1: float, l: float, b: float = B) -> Tuple[float, float]:
    """同板前→后：给定 θ1 和 l，返回 (r2=bθ2, θ2)"""
    delta = solve_delta_same_board(theta1, l, b)
    theta2 = theta1 + delta
    r2 = b * theta2
    return r2, theta2


# --------------- 在时刻 t 计算 223+? 个把手 ---------------
def handles_at_time(t: float) -> Tuple[np.ndarray, np.ndarray, List[str]]:
    """
    返回：
      X, Y: shape (N,) —— 所有“前把手 + 尾部把手”的笛卡尔坐标
      names: 名称列表（与你当前程序一致，不改动）
    迭代（保留你当前口径）：
      H_front --(l_HEAD)--> B1_front
               --(l_BODY)--> B2_front
               ...
               --(l_BODY)--> B221_front
      然后：221 前 -> 221 后（l_BODY） -> 尾后（你当前设定距离 = 1.65）
    """
    # 龙头前把手
    r, th = head_polar_at_time(t)
    xs = [r * math.cos(th)]
    ys = [r * math.sin(th)]
    names = ["龙头-前把手"]

    # 第1节龙身前把手 = “龙头同板后把手”
    r_f, th_f = step_same_board(th, l_HEAD)
    xs.append(r_f * math.cos(th_f))
    ys.append(r_f * math.sin(th_f))
    names.append("第1节龙身-前把手")

    # 继续 2..221 节（每次加 l_BODY）
    th_cur = th_f
    for m in range(2, 222):
        r_f, th_f = step_same_board(th_cur, l_BODY)
        xs.append(r_f * math.cos(th_f))
        ys.append(r_f * math.sin(th_f))
        names.append(f"第{m}节龙身-前把手")
        th_cur = th_f

    # —— 按你当前定义仅重算尾部（不改动你的设定：1.65） ——
    # 现在 th_cur 是“第221节 前把手”的 θ
    rh_221_back, th_221_back = step_same_board(th_cur, l_BODY)    # 221 前 -> 221 后
    xs.append(rh_221_back * math.cos(th_221_back))
    ys.append(rh_221_back * math.sin(th_221_back))
    names.append("龙尾-前把手")  # 保留你当前的列头

    L_221BACK_to_TAILREAR = 1.65  # 你现在的口径（我不改）
    r_tail, th_tail = step_same_board(th_221_back, L_221BACK_to_TAILREAR)
    xs.append(r_tail * math.cos(th_tail))
    ys.append(r_tail * math.sin(th_tail))
    names.append("龙尾-后把手")

    return np.array(xs), np.array(ys), names


# --------------- 数值求导：由位置算速度（vx, vy, speed） ---------------
def numerical_derivative_series(arr_2d: np.ndarray, dt: float) -> np.ndarray:
    """
    对时间序列矩阵 arr_2d 逐列数值求导：
      - 端点：前/后向差分
      - 内点：中心差分
    arr_2d 形状：(T, N)
    返回同形矩阵 (T, N)
    """
    T, N = arr_2d.shape
    der = np.zeros_like(arr_2d, dtype=float)
    if T >= 2:
        der[0, :]  = (arr_2d[1, :] - arr_2d[0, :]) / dt
        der[-1, :] = (arr_2d[-1, :] - arr_2d[-2, :]) / dt
    if T >= 3:
        der[1:-1, :] = (arr_2d[2:, :] - arr_2d[:-2, :]) / (2.0 * dt)
    return der

# ==== 题面参数（几何） ====
W_BOARD = 0.30        # 板宽 (m) —— 题面：30 cm
HALF_W  = W_BOARD/2.0 # 胶囊半径 r
EXT_END = 0.275       # 孔心到端头的延伸 (m) —— 题面：27.5 cm

# ==== 向量工具 ====
import numpy as np

def _norm(v: np.ndarray, eps: float = 1e-12) -> float:
    n = float(np.hypot(v[0], v[1]))
    return n if n > eps else eps

def _unit(v: np.ndarray) -> np.ndarray:
    n = _norm(v)
    return v / n

# ==== 用“把手序列”构造 223 块板的中心线段（两端已按 0.275m 外延） ====
def build_board_segments(x: np.ndarray, y: np.ndarray):
    """
    输入:
        x, y : shape (N,) —— handles_at_time 返回的所有把手(含龙尾后把手)
                其顺序恰好是每两相邻把手 (i, i+1) 构成同一块板
                => 共 N-1 块板（应为 223）
    返回:
        segs: list of (A, B)，A/B 为 np.array([x, y])，是“整块板”的中心线端点
              端点 = [前孔心]向前外延0.275m & [后孔心]向后外延0.275m
    """
    N = len(x)
    segs = []
    for i in range(N - 1):
        p_front = np.array([x[i],   y[i]  ], dtype=float)
        p_back  = np.array([x[i+1], y[i+1]], dtype=float)
        d = _unit(p_back - p_front)            # 沿“同板前后把手”方向的单位向量
        A = p_front - d * EXT_END              # 前端头
        B = p_back  + d * EXT_END              # 后端头
        segs.append((A, B))
    return segs  # 长度应为 223



# ========== 用线段构造“平端矩形板”四个顶点 ==========
def rect_corners_from_segment(A: np.ndarray, B: np.ndarray, width: float) -> np.ndarray:
    """
    输入：A,B 为整块板中心线段两端点（你已按 0.275 m 外延过），width=板宽W
    输出：4x2 顶点（顺时针），板端为“平端”，不是圆头
    """
    d = _unit(B - A)                       # 板方向
    n = np.array([-d[1], d[0]], float)     # 板法向
    hw = width / 2.0
    # 顺时针：A侧上 -> B侧上 -> B侧下 -> A侧下
    return np.vstack([A + n*hw, B + n*hw, B - n*hw, A - n*hw])

# ========== 多边形在轴上的投影区间 ==========
def _project_interval(axis: np.ndarray, pts: np.ndarray) -> Tuple[float, float]:
    a = axis / (_norm(axis) + 1e-15)
    s = pts @ a
    return float(s.min()), float(s.max())

# ========== 矩形对的“带符号安全裕度” (SAT) ==========
def rect_pair_margin_SAT(c1: np.ndarray, c2: np.ndarray) -> float:
    """
    返回一对矩形的“安全裕度”：
      >0 代表分离（取所有轴上的“最大间隙”）
      =0 代表恰好接触
      <0 代表重叠（负号为“最小穿透深度”）
    采用四条轴：rect1 的 (d1,n1)，rect2 的 (d2,n2)
    """
    # 由顶点恢复两主轴
    d1 = c1[1] - c1[0]; n1 = c1[3] - c1[0]
    d2 = c2[1] - c2[0]; n2 = c2[3] - c2[0]
    axes = [d1, n1, d2, n2]

    best_gap = -1.0e18      # 分离时的“最大间隙”
    min_overlap = +1.0e18   # 相交时的“最小重叠”

    for ax in axes:
        a1, b1 = _project_interval(ax, c1)
        a2, b2 = _project_interval(ax, c2)
        overlap = min(b1, b2) - max(a1, a2)
        if overlap < 0.0:                 # 分离
            gap = -overlap
            if gap > best_gap:
                best_gap = gap
        else:                              # 相交（或接触）
            if overlap < min_overlap:
                min_overlap = overlap

    if best_gap > 0.0:
        return best_gap          # 分离，正数 = 间隙
    return -min_overlap          # 重叠/接触，负/零 = -穿透(或0)

# ========== 用矩形-SAT计算“全局最小安全裕度” μ(t) ==========
def min_clearance_for_time(t: float) -> float:
    """
    μ(t) = min_{|i-j|>=2} margin_SAT(R_i, R_j)
      - |i-j|=1（相邻同把手连接）忽略
      - margin_SAT>0 安全；=0 刚接触；<0 碰撞
    """
    X, Y, _ = handles_at_time(t)
    segs = build_board_segments(X, Y)     # 223 段（整板中心线）
    rects = [rect_corners_from_segment(A, B, W_BOARD) for (A, B) in segs]

    mu = +1e9
    m = len(rects)
    for i in range(m):
        ri = rects[i]
        for j in range(i+2, m):           # 跳过相邻
            rj = rects[j]
            margin = rect_pair_margin_SAT(ri, rj)
            if margin < mu:
                mu = margin
            if mu < 0.0:                  # 发现碰撞可早停
                return mu
    return mu

# ==== 搜索“最晚不碰撞时刻” t_end ====
def find_terminal_time(t_max: float = 300.0,
                       coarse_dt: float = 1.0,
                       bisect_tol: float = 1e-3) -> float:
    """
    先用粗步长扫描，定位首次从安全(μ>=0)变为不安全(μ<0)的 [t_lo, t_hi]，
    再在区间内二分 μ(t)=0 的根，得到 t_end（最后安全时刻）。
    """
    # 粗扫描
    t_grid = np.arange(0.0, t_max + 1e-12, coarse_dt, dtype=float)
    mu_prev = None
    t_lo = 0.0
    for t in t_grid:
        mu = min_clearance_for_time(t)
        if mu_prev is None:
            mu_prev = mu
            t_lo = t
            continue
        if mu_prev >= 0.0 and mu < 0.0:
            t_hi = t
            break
        mu_prev = mu
        t_lo = t
    else:
        # 扫描到 t_max 仍安全：按题意就是“还能继续盘入”，返回 t_max
        return float(t_max)

    # 二分细化
    lo, hi = t_lo, t_hi
    for _ in range(60):
        mid = 0.5 * (lo + hi)
        mu_mid = min_clearance_for_time(mid)
        if mu_mid >= 0.0:
            lo = mid
        else:
            hi = mid
        if (hi - lo) < bisect_tol:
            break
    return float(lo)  # 最晚安全时刻

from openpyxl import load_workbook

# ---------- 名称归一化：把程序里的名称映射到模板的名称 ----------
def _name_for_template(nm: str) -> str:
    """
    程序里常见：
      - "龙头-前把手"    -> 模板："龙头"
      - "第k节龙身-前把手" -> 模板："第k节龙身"
      - "龙尾-前把手"    -> 模板："龙尾"
      - "龙尾-后把手"    -> 模板："龙尾（后）"
    同时兼容本来就是模板名字的情况（直接返回）。
    """
    s = str(nm).strip()
    if s.startswith("龙头"):
        return "龙头"
    if s.startswith("第") and "节龙身" in s:
        return s.split("-")[0]
    if s.startswith("龙尾-前把手"):
        return "龙尾"
    if s.startswith("龙尾-后把手"):
        return "龙尾（后）"
    return s  # 已经是模板名时直接返回

def solve_q2_and_fill_template(
    template_path: str = "D:/MathModeling/2024国赛A题/附件/result2.xlsx",
    vel_dt: float = 0.1,
    coarse_dt: float = 1.0,
    bisect_tol: float = 1e-3,
):
    """
    第二问主函数（不改计算思路，只负责把结果写进你给的 Sheet1）：
      1) 调用你已有的 t_end 求解方法（find_terminal_time）
      2) 在 t_end 处：计算所有把手 (x,y)
      3) 用中心差分在 t_end 近邻估算速度合量 speed，龙头速度强制=1
      4) 打开模板 result2.xlsx（Sheet1），按“名称”列精准写入三列数值
    参数：
      - template_path: 你提供的模板文件路径（会原地覆盖内容）
      - vel_dt: 计算速度的中心差分步长（s），默认 0.1
      - coarse_dt, bisect_tol: 传给你已有的 t_end 搜索函数
    """
    # 1) 求 t_end（沿用你现有的计算思路/函数）
    t_end = find_terminal_time(t_max=500, coarse_dt=coarse_dt, bisect_tol=bisect_tol)
    print(f"[Q2] t_end = {t_end:.6f} s")

    # 2) t_end 位置
    X0, Y0, names0 = handles_at_time(t_end)

    # 3) 速度（取最晚时刻的“端点后向差分”，仍然复用你的 numerical_derivative_series）
    t_prev = max(T0, t_end - DT)            # 与你的时间网格保持同样步长 DT
    X_prev, Y_prev, _ = handles_at_time(t_prev)

    # 组装 2×N，按列求导：第一行对应 t_prev，第二行对应 t_end
    X_pair = np.vstack([X_prev, X0])        # 形状 (2, N)
    Y_pair = np.vstack([Y_prev, Y0])        # 形状 (2, N)

    VX_pair = numerical_derivative_series(X_pair, DT)
    VY_pair = numerical_derivative_series(Y_pair, DT)

    # 最后一行就是 t_end 的速度（后向差分）
    SPEED = np.hypot(VX_pair[-1, :], VY_pair[-1, :])
    SPEED[0] = 1.0   # 龙头速度按题面强制为 1 m/s

    # 4) 建立 “模板名称 -> 程序索引” 的映射
    name_to_idx = {}
    for j, nm in enumerate(names0):
        name_to_idx[_name_for_template(nm)] = j

    # 5) 打开模板并写入
    wb = load_workbook(template_path)
    if "Sheet1" not in wb.sheetnames:
        raise ValueError("模板中未找到 Sheet1，请检查文件。")
    ws = wb["Sheet1"]

    # 假设首行是表头：A1=名称, B1=横坐标x (m), C1=纵坐标y (m), D1=速度 (m/s)
    # 从第2行开始逐行填充
    # 同时兼容表头名字略有出入（按列位置来写更稳妥）
    col_name = 1  # A
    col_x    = 2  # B
    col_y    = 3  # C
    col_spd  = 4  # D

    # 找到最后一行：直到 A 列为空视为结束（或用 ws.max_row）
    max_row = ws.max_row
    # 若模板中 A 列到底都填了名称，这里直接用 max_row 会更方便
    for r in range(2, max_row + 1):
        nm_tpl = ws.cell(row=r, column=col_name).value
        if nm_tpl is None:
            continue
        nm_tpl = str(nm_tpl).strip()
        if nm_tpl in name_to_idx:
            j = name_to_idx[nm_tpl]
            ws.cell(row=r, column=col_x  , value=round(float(X0[j]), 6))
            ws.cell(row=r, column=col_y  , value=round(float(Y0[j]), 6))
            ws.cell(row=r, column=col_spd, value=round(float(SPEED[j]), 6))
        else:
            # 模板有该名称，但程序没找到对应（比如拼写差异）
            # 可打印提醒方便排查
            # print(f"[WARN] 模板名称未匹配：{nm_tpl}")
            pass

    wb.save(template_path)
    print(f"[Q2] 已填入模板：{template_path}")

solve_q2_and_fill_template()


import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon

# ---------- 基础几何：由中心点 + 朝向向量 构造矩形四角 ----------
def _rect_corners(center_xy, axis_vec, L, W):
    """
    输入：
      center_xy: 矩形中心 (x, y)
      axis_vec : 朝向向量（指向“后把手-前把手”）
      L, W     : 矩形长度、宽度
    输出：四个角点坐标（顺时针）
    理论：û = axis/||axis||，n = R90(û)，角点 = C ± (L/2)*û ± (W/2)*n
    """
    cx, cy = float(center_xy[0]), float(center_xy[1])
    ax, ay = float(axis_vec[0]),  float(axis_vec[1])
    norm = (ax**2 + ay**2) ** 0.5
    if norm < 1e-12:
        ux, uy = 1.0, 0.0
    else:
        ux, uy = ax / norm, ay / norm
    nx, ny = -uy, ux
    halfL, halfW = 0.5 * L, 0.5 * W
    p1 = (cx + halfL*ux + halfW*nx, cy + halfL*uy + halfW*ny)
    p2 = (cx + halfL*ux - halfW*nx, cy + halfL*uy - halfW*ny)
    p3 = (cx - halfL*ux - halfW*nx, cy - halfL*uy - halfW*ny)
    p4 = (cx - halfL*ux + halfW*nx, cy - halfL*uy + halfW*ny)
    return [p1, p2, p3, p4]

# ---------- 生成“板”的前/后把手（最终时刻） ----------
def boards_at_time(t: float):
    """
    返回 items 列表，每项：
      {
        'name': '龙头板' / '第k节龙身' / '龙尾板',
        'F': 前把手 (x,y),
        'B': 后把手 (x,y),
        'L': 板长（L_HEAD/L_BODY/L_TAIL）
      }
    规则：同板前后把手中心距 l = L - 0.55；两把手都在同一等距螺线 r=bθ 上。
    """
    items = []

    # —— 龙头板 ——
    r0, th0 = head_polar_at_time(t)                                  # 龙头前把手
    F_head = np.array([r0*np.cos(th0), r0*np.sin(th0)], dtype=float)
    r2, th2 = step_same_board(th0, l_HEAD)                           # 龙头同板后把手
    B_head = np.array([r2*np.cos(th2), r2*np.sin(th2)], dtype=float)
    items.append({'name': '龙头板', 'F': F_head, 'B': B_head, 'L': L_HEAD})

    # —— 1..221 节龙身 ——
    th_cur = th2  # 第1节龙身-前把手 = 龙头后把手
    for k in range(1, 222):
        rf = B * th_cur
        Fk = np.array([rf*np.cos(th_cur), rf*np.sin(th_cur)], dtype=float)  # 第k节前把手
        rb, thb = step_same_board(th_cur, l_BODY)                           # 第k节后把手
        Bk = np.array([rb*np.cos(thb), rb*np.sin(thb)], dtype=float)
        items.append({'name': f'第{k}节龙身', 'F': Fk, 'B': Bk, 'L': L_BODY})
        th_cur = thb  # 下一节的前把手角

    # —— 龙尾板（前把手=上一节后把手；再走 l_TAIL 得尾后） ——
    rf_tail = B * th_cur
    F_tail = np.array([rf_tail*np.cos(th_cur), rf_tail*np.sin(th_cur)], dtype=float)
    rb_tail, th_tail = step_same_board(th_cur, l_TAIL)               # l_TAIL = L_TAIL - 0.55 (=1.65)
    B_tail = np.array([rb_tail*np.cos(th_tail), rb_tail*np.sin(th_tail)], dtype=float)
    items.append({'name': '龙尾板', 'F': F_tail, 'B': B_tail, 'L': L_TAIL})

    return items

# ---------- 仅画最终时刻：矩形 + 完整等距螺线(从原点到最大半径) ----------
def plot_final_layout(t: float = 412.473633,
                      board_width: float = 0.3,
                      color: str = "#1f77b4",    # 统一颜色
                      draw_spiral: bool = True,
                      draw_handles: bool = True,
                      save_png: bool = True,
                      png_path: str = "D:/MathModeling/2024国赛A题/附件/Problem2可视化结果.png"):
    """
    可视化最终时刻布局：
      - 船身/板：矩形（统一颜色，半透明）
      - 把手散点：同色（可关闭）
      - 等距螺线：从原点 (0,0) 开始，完整画到覆盖全部把手（θ ∈ [0, θ_max]）
    关键点：
      * 完整螺线：θ 从 0 到 θ_max，其中 θ_max = max_i ( r_i / B )，r_i 为所有把手半径
      * θ=0 => r=0 => 就是原点，严格画到 (0,0)
    """
    items = boards_at_time(t)

    fig, ax = plt.subplots(figsize=(9, 9))

    # —— 完整等距螺线：θ ∈ [0, θ_max]，保证到 (0,0) —— #
    if draw_spiral:
        # 汇总全部把手的半径 r_i，求 θ_max = max(r_i)/B
        radii = []
        for it in items:
            radii.append(float(np.hypot(it['F'][0], it['F'][1])))
            radii.append(float(np.hypot(it['B'][0], it['B'][1])))
        r_max = max(radii) if radii else 0.0
        theta_max = r_max / B

        # 采样点数随 θ_max 自适应，保证平滑；首点强制 (0,0)
        n_pts = max(1200, int(400 * theta_max))
        thetas = np.linspace(0.0, max(theta_max, 0.0), n_pts)
        rs = B * thetas
        xs, ys = rs * np.cos(thetas), rs * np.sin(thetas)
        xs[0], ys[0] = 0.0, 0.0  # 明确首点为原点
        ax.plot(xs, ys, ls='--', lw=1.4, color=color, alpha=0.9, label='等距螺线 r=bθ ')

    # —— 画所有板的矩形（统一颜色） —— #
    all_pts = []
    for it in items:
        F = it['F']; Bp = it['B']; L = it['L']
        C = 0.5 * (F + Bp)
        axis = (Bp - F)
        corners = _rect_corners(C, axis, L=L, W=board_width)
        poly = Polygon(corners, closed=True,
                       facecolor=color, edgecolor=color, lw=0.6, alpha=0.35)
        ax.add_patch(poly)
        all_pts.extend(corners)

    # —— 可选：把手散点（统一颜色） —— #
    if draw_handles:
        Fx = [it['F'][0] for it in items]; Fy = [it['F'][1] for it in items]
        Bx = [it['B'][0] for it in items]; By = [it['B'][1] for it in items]
        ax.scatter(Fx, Fy, s=12, color=color, label='前把手', zorder=5)
        ax.scatter(Bx, By, s=12, color=color, marker='s', label='后把手', zorder=5)
        # 龙头/龙尾强调（同色不同形状）
        ax.scatter([items[0]['F'][0]],  [items[0]['F'][1]],  s=40, color=color, marker='*', label='龙头-前把手', zorder=6)
        ax.scatter([items[-1]['B'][0]], [items[-1]['B'][1]], s=40, color=color, marker='^', label='龙尾-后把手', zorder=6)

    # —— 视野设置 —— #
    if all_pts:
        all_pts = np.array(all_pts)
        pad = 0.6
        xmin, ymin = all_pts.min(axis=0) - pad
        xmax, ymax = all_pts.max(axis=0) + pad
        ax.set_xlim(xmin, xmax); ax.set_ylim(ymin, ymax)

    ax.set_aspect('equal', adjustable='box')
    ax.grid(True, ls='--', alpha=0.25)
    ax.set_title(f"最终时刻 t = {t:.0f} s 的布局")
    ax.set_xlabel("x / m"); ax.set_ylabel("y / m")
    ax.legend(loc='best')
    plt.tight_layout()

    if save_png:
        plt.savefig(png_path, dpi=300)
        print(f"[OK] 已保存：{png_path}")

    plt.show()

plot_final_layout()
