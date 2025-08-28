import math
from typing import Tuple, List
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.patches import Polygon
plt.rcParams['font.family'] = 'SimHei'
plt.rcParams['axes.unicode_minus'] = False
# --------------- 基本参数 ---------------
PITCH = 0.55                        # 螺距 p（m）
B = PITCH / (2.0 * math.pi)         # b = p/(2π)
THETA0 = 2.0 * math.pi * 16         # t=0 时龙头角（第 16 圈 A 点）
V = 1.0                             # 龙头弧长速度（m/s）

# 板长（米）与“同板前后把手中心距离” l = L - 0.55
L_HEAD, L_BODY, L_TAIL = 3.41, 2.20, 2.20
D_HOLE = 0.55
l_HEAD = L_HEAD - D_HOLE            # 2.86
l_BODY = L_BODY - D_HOLE            # 1.65
l_TAIL = L_TAIL - D_HOLE            # 1.65

# 时间与输出
T0, T1, DT = 0.0, 300.0, 1.0
PAPER_TIMES = [0, 60, 120, 180, 240, 300]
OUT_XLSX = "D:/MathModeling/2024国赛A题/附件/result1.xlsx"


# --------------- 弧长原函数与反解 ---------------
def F_theta(theta: float) -> float:
    """F(θ) = 0.5*( θ*sqrt(1+θ^2) + asinh(θ) )"""
    return 0.5 * (theta * math.sqrt(1.0 + theta * theta) + math.asinh(theta))

def inv_F(Fv: float, max_iter: int = 40, tol: float = 1e-13) -> float:
    """牛顿法反解 θ，使 F(θ) = Fv（θ≥0，F 单调增）"""
    theta = math.sqrt(2.0 * Fv) if Fv > 1.0 else Fv
    for _ in range(max_iter):
        f = F_theta(theta) - Fv
        df = math.sqrt(1.0 + theta * theta)
        theta_new = theta - f / df
        if theta_new < 0.0:
            theta_new = 0.0
        if abs(theta_new - theta) < tol:
            theta = theta_new
            break
        theta = theta_new
    return theta

def head_polar_at_time(t: float) -> Tuple[float, float]:
    """给定 t，求龙头极坐标 (r, θ)。顺时针盘入：F(θ)=F(θ0)-t/b"""
    F0 = F_theta(THETA0)
    theta = inv_F(F0 - t / B)      # v=1
    r = B * theta
    return r, theta


# --------------- 求同板 Δ（最近物理解，Δ∈(0, π)） ---------------
def solve_delta_same_board(theta1: float, l: float, b: float = B) -> float:
    """
    已知同一条螺线上的“前把手”角 θ1（r1 = bθ1），求同板“后把手”使两点直线距离为 l。
    设 Δ = θ2 - θ1 > 0，r2 = b(θ1 + Δ)，方程：
      g(Δ) = r1^2 + r2^2 - 2 r1 r2 cosΔ - l^2 = 0
    取最近的物理解 Δ ∈ (0, π)。先牛顿（带投影/阻尼），失败退二分。
    """
    r1 = b * theta1

    def g(delta: float) -> float:
        r2 = b * (theta1 + delta)
        return r1*r1 + r2*r2 - 2.0*r1*r2*math.cos(delta) - l*l

    def gprime(delta: float) -> float:
        # r2 = b(θ1+Δ)
        r2 = b * (theta1 + delta)
        # dg/dΔ = 2 r2 b - 2 r1 b cosΔ + 2 r1 r2 sinΔ
        return 2.0*r2*b - 2.0*r1*b*math.cos(delta) + 2.0*r1*r2*math.sin(delta)

    # 牛顿初值：小角近似 Δ ≈ l / max(r1, 1e-9)
    delta = min(max(l / max(r1, 1e-9), 1e-10), math.pi/2)
    ok = False
    for _ in range(20):
        val = g(delta)
        der = gprime(delta)
        if abs(der) < 1e-14:
            break
        # 阻尼牛顿 + 投影到 (0, π)
        cand = delta - val/der
        if not (0.0 < cand < math.pi):
            cand = 0.5*(delta + max(1e-10, min(cand, math.pi-1e-10)))
        delta = cand
        if abs(val) < 1e-12:
            ok = True
            break
    if ok:
        return delta

    # 二分：在 (0, π) 内找根
    low, high = 1e-12, math.pi - 1e-12
    gl, gh = g(low), g(high)   # g(0+)≈-l^2<0, g(π-)通常>0
    if gl*gh > 0:
        # 极端场景：略扩上界再试
        high = min(2.0*math.pi, high + 1.0)
        gh = g(high)
        if gl*gh > 0:
            return delta  # 退回牛顿最后值
    for _ in range(80):
        mid = 0.5*(low + high)
        gm = g(mid)
        if gm == 0.0 or (high - low) < 1e-12:
            return mid
        if gl*gm <= 0.0:
            high, gh = mid, gm
        else:
            low, gl = mid, gm
    return 0.5*(low + high)


def step_same_board(theta1: float, l: float, b: float = B) -> Tuple[float, float]:
    """同板前→后：给定 θ1 和 l，返回 (r2=bθ2, θ2)"""
    delta = solve_delta_same_board(theta1, l, b)
    theta2 = theta1 + delta
    r2 = b * theta2
    return r2, theta2


# --------------- 在时刻 t 计算 223+? 个把手 ---------------
def handles_at_time(t: float) -> Tuple[np.ndarray, np.ndarray, List[str]]:
    """
    返回：
      X, Y: shape (N,) —— 所有“前把手 + 尾部把手”的笛卡尔坐标
      names: 名称列表（与你当前程序一致，不改动）
    迭代（保留你当前口径）：
      H_front --(l_HEAD)--> B1_front
               --(l_BODY)--> B2_front
               ...
               --(l_BODY)--> B221_front
      然后：221 前 -> 221 后（l_BODY） -> 尾后（你当前设定距离 = 1.65）
    """
    # 龙头前把手
    r, th = head_polar_at_time(t)
    xs = [r * math.cos(th)]
    ys = [r * math.sin(th)]
    names = ["龙头-前把手"]

    # 第1节龙身前把手 = “龙头同板后把手”
    r_f, th_f = step_same_board(th, l_HEAD)
    xs.append(r_f * math.cos(th_f))
    ys.append(r_f * math.sin(th_f))
    names.append("第1节龙身-前把手")

    # 继续 2..221 节（每次加 l_BODY）
    th_cur = th_f
    for m in range(2, 222):
        r_f, th_f = step_same_board(th_cur, l_BODY)
        xs.append(r_f * math.cos(th_f))
        ys.append(r_f * math.sin(th_f))
        names.append(f"第{m}节龙身-前把手")
        th_cur = th_f

    # —— 按你当前定义仅重算尾部（不改动你的设定：1.65） ——
    # 现在 th_cur 是“第221节 前把手”的 θ
    rh_221_back, th_221_back = step_same_board(th_cur, l_BODY)    # 221 前 -> 221 后
    xs.append(rh_221_back * math.cos(th_221_back))
    ys.append(rh_221_back * math.sin(th_221_back))
    names.append("龙尾-前把手")  # 保留你当前的列头

    L_221BACK_to_TAILREAR = 1.65  # 你现在的口径（我不改）
    r_tail, th_tail = step_same_board(th_221_back, L_221BACK_to_TAILREAR)
    xs.append(r_tail * math.cos(th_tail))
    ys.append(r_tail * math.sin(th_tail))
    names.append("龙尾-后把手")

    return np.array(xs), np.array(ys), names


# --------------- 数值求导：由位置算速度（vx, vy, speed） ---------------
def numerical_derivative_series(arr_2d: np.ndarray, dt: float) -> np.ndarray:
    """
    对时间序列矩阵 arr_2d 逐列数值求导：
      - 端点：前/后向差分
      - 内点：中心差分
    arr_2d 形状：(T, N)
    返回同形矩阵 (T, N)
    """
    T, N = arr_2d.shape
    der = np.zeros_like(arr_2d, dtype=float)
    if T >= 2:
        der[0, :]  = (arr_2d[1, :] - arr_2d[0, :]) / dt
        der[-1, :] = (arr_2d[-1, :] - arr_2d[-2, :]) / dt
    if T >= 3:
        der[1:-1, :] = (arr_2d[2:, :] - arr_2d[:-2, :]) / (2.0 * dt)
    return der


from openpyxl import load_workbook

def export_positions(path: str = OUT_XLSX):
    """
    按你上传的模板 result1.xlsx 进行写入：
      - Sheet『位置』：列为时间 0..300 s，行依次为 龙头x、龙头y、 第1~221节龙身x/y、龙尾x/y、龙尾（后）x/y
      - Sheet『速度』：同样的行顺序，但每行仅填“合速度 (m/s)”，不写 vx/vy
    重要约束：
      - 龙头速度恒定为 1 m/s（题面要求），因此这里直接强制写 1.000000，避免数值差分的细小误差
      - 其它把手的速度用中心差分(端点用前/后向差分)对位置进行数值求导得到
      - 结果统一保留 6 位小数
      - 不创建任何新工作表，只更新模板中的『位置』『速度』两张表
    """
    # --------- 时间网格 ---------
    ts = np.arange(T0, T1 + 1e-12, DT, dtype=float)   # 0..300 步长1，共 301 列
    T = ts.size

    # --------- 全时刻计算位置（X(t), Y(t)）---------
    X_list, Y_list = [], []
    x0, y0, names0 = handles_at_time(ts[0])
    N = len(names0)   # 应为 224（龙头、221节身、龙尾前、龙尾后）
    for t in ts:
        xi, yi, _ = handles_at_time(t)
        X_list.append(xi); Y_list.append(yi)
    X = np.vstack(X_list)  # (T, N)
    Y = np.vstack(Y_list)  # (T, N)

    # --------- 数值求导得到合速度 ---------
    VX = numerical_derivative_series(X, DT)
    VY = numerical_derivative_series(Y, DT)
    SPEED = np.sqrt(VX*VX + VY*VY)
    SPEED[:, 0] = 1.0  # 龙头速度按题面强制设为 1 m/s（避免数值误差）

    # --------- 打开模板，定位两张表 ---------
    wb = load_workbook(path)
    ws_pos = wb["位置"]
    ws_spd = wb["速度"]

    # 写入『位置』
    for ti in range(T):
        col = 2 + ti  # 第2列对应 0 s
        # 写所有把手
        for j in range(N):
            base_row = 2 + 2*j
            # x行
            ws_pos.cell(row=base_row,   column=col, value=round(float(X[ti, j]), 6))
            # y行
            ws_pos.cell(row=base_row+1, column=col, value=round(float(Y[ti, j]), 6))

    # 写入『速度』（仅合速度，单位 m/s）
    for ti in range(T):
        col = 2 + ti
        for j in range(N):
            ws_spd.cell(row=2 + j, column=col, value=round(float(SPEED[ti, j]), 6))

    # 保存到同一路径（覆盖写入数值，保留模板格式）
    wb.save(path)

export_positions(OUT_XLSX)

# ---------- 基础几何：由中心点 + 朝向向量 构造矩形四角 ----------
def _rect_corners(center_xy, axis_vec, L, W):
    """
    输入：
      center_xy: 矩形中心 (x, y)
      axis_vec : 朝向向量（指向“后把手-前把手”）
      L, W     : 矩形长度、宽度
    输出：四个角点坐标（顺时针）
    理论：û = axis/||axis||，n = R90(û)，角点 = C ± (L/2)*û ± (W/2)*n
    """
    cx, cy = float(center_xy[0]), float(center_xy[1])
    ax, ay = float(axis_vec[0]),  float(axis_vec[1])
    norm = (ax**2 + ay**2) ** 0.5
    if norm < 1e-12:
        ux, uy = 1.0, 0.0
    else:
        ux, uy = ax / norm, ay / norm
    nx, ny = -uy, ux
    halfL, halfW = 0.5 * L, 0.5 * W
    p1 = (cx + halfL*ux + halfW*nx, cy + halfL*uy + halfW*ny)
    p2 = (cx + halfL*ux - halfW*nx, cy + halfL*uy - halfW*ny)
    p3 = (cx - halfL*ux - halfW*nx, cy - halfL*uy - halfW*ny)
    p4 = (cx - halfL*ux + halfW*nx, cy - halfL*uy + halfW*ny)
    return [p1, p2, p3, p4]

# ---------- 生成“板”的前/后把手（最终时刻） ----------
def boards_at_time(t: float):
    """
    返回 items 列表，每项：
      {
        'name': '龙头板' / '第k节龙身' / '龙尾板',
        'F': 前把手 (x,y),
        'B': 后把手 (x,y),
        'L': 板长（L_HEAD/L_BODY/L_TAIL）
      }
    规则：同板前后把手中心距 l = L - 0.55；两把手都在同一等距螺线 r=bθ 上。
    """
    items = []

    # —— 龙头板 ——
    r0, th0 = head_polar_at_time(t)                                  # 龙头前把手
    F_head = np.array([r0*np.cos(th0), r0*np.sin(th0)], dtype=float)
    r2, th2 = step_same_board(th0, l_HEAD)                           # 龙头同板后把手
    B_head = np.array([r2*np.cos(th2), r2*np.sin(th2)], dtype=float)
    items.append({'name': '龙头板', 'F': F_head, 'B': B_head, 'L': L_HEAD})

    # —— 1..221 节龙身 ——
    th_cur = th2  # 第1节龙身-前把手 = 龙头后把手
    for k in range(1, 222):
        rf = B * th_cur
        Fk = np.array([rf*np.cos(th_cur), rf*np.sin(th_cur)], dtype=float)  # 第k节前把手
        rb, thb = step_same_board(th_cur, l_BODY)                           # 第k节后把手
        Bk = np.array([rb*np.cos(thb), rb*np.sin(thb)], dtype=float)
        items.append({'name': f'第{k}节龙身', 'F': Fk, 'B': Bk, 'L': L_BODY})
        th_cur = thb  # 下一节的前把手角

    # —— 龙尾板（前把手=上一节后把手；再走 l_TAIL 得尾后） ——
    rf_tail = B * th_cur
    F_tail = np.array([rf_tail*np.cos(th_cur), rf_tail*np.sin(th_cur)], dtype=float)
    rb_tail, th_tail = step_same_board(th_cur, l_TAIL)               # l_TAIL = L_TAIL - 0.55 (=1.65)
    B_tail = np.array([rb_tail*np.cos(th_tail), rb_tail*np.sin(th_tail)], dtype=float)
    items.append({'name': '龙尾板', 'F': F_tail, 'B': B_tail, 'L': L_TAIL})

    return items

# ---------- 仅画最终时刻：矩形 + 完整等距螺线(从原点到最大半径) ----------
def plot_final_layout(t: float = T1,
                      board_width: float = 0.3,
                      color: str = "#1f77b4",    # 统一颜色
                      draw_spiral: bool = True,
                      draw_handles: bool = True,
                      save_png: bool = True,
                      png_path: str = "D:/MathModeling/2024国赛A题/附件/Problem1可视化结果.png"):
    """
    可视化最终时刻布局：
      - 船身/板：矩形（统一颜色，半透明）
      - 把手散点：同色（可关闭）
      - 等距螺线：从原点 (0,0) 开始，完整画到覆盖全部把手（θ ∈ [0, θ_max]）
    关键点：
      * 完整螺线：θ 从 0 到 θ_max，其中 θ_max = max_i ( r_i / B )，r_i 为所有把手半径
      * θ=0 => r=0 => 就是原点，严格画到 (0,0)
    """
    items = boards_at_time(t)

    fig, ax = plt.subplots(figsize=(9, 9))

    # —— 完整等距螺线：θ ∈ [0, θ_max]，保证到 (0,0) —— #
    if draw_spiral:
        # 汇总全部把手的半径 r_i，求 θ_max = max(r_i)/B
        radii = []
        for it in items:
            radii.append(float(np.hypot(it['F'][0], it['F'][1])))
            radii.append(float(np.hypot(it['B'][0], it['B'][1])))
        r_max = max(radii) if radii else 0.0
        theta_max = r_max / B

        # 采样点数随 θ_max 自适应，保证平滑；首点强制 (0,0)
        n_pts = max(1200, int(400 * theta_max))
        thetas = np.linspace(0.0, max(theta_max, 0.0), n_pts)
        rs = B * thetas
        xs, ys = rs * np.cos(thetas), rs * np.sin(thetas)
        xs[0], ys[0] = 0.0, 0.0  # 明确首点为原点
        ax.plot(xs, ys, ls='--', lw=1.4, color=color, alpha=0.9, label='等距螺线 r=bθ ')

    # —— 画所有板的矩形（统一颜色） —— #
    all_pts = []
    for it in items:
        F = it['F']; Bp = it['B']; L = it['L']
        C = 0.5 * (F + Bp)
        axis = (Bp - F)
        corners = _rect_corners(C, axis, L=L, W=board_width)
        poly = Polygon(corners, closed=True,
                       facecolor=color, edgecolor=color, lw=0.6, alpha=0.35)
        ax.add_patch(poly)
        all_pts.extend(corners)

    # —— 可选：把手散点（统一颜色） —— #
    if draw_handles:
        Fx = [it['F'][0] for it in items]; Fy = [it['F'][1] for it in items]
        Bx = [it['B'][0] for it in items]; By = [it['B'][1] for it in items]
        ax.scatter(Fx, Fy, s=12, color=color, label='前把手', zorder=5)
        ax.scatter(Bx, By, s=12, color=color, marker='s', label='后把手', zorder=5)
        # 龙头/龙尾强调（同色不同形状）
        ax.scatter([items[0]['F'][0]],  [items[0]['F'][1]],  s=40, color=color, marker='*', label='龙头-前把手', zorder=6)
        ax.scatter([items[-1]['B'][0]], [items[-1]['B'][1]], s=40, color=color, marker='^', label='龙尾-后把手', zorder=6)

    # —— 视野设置 —— #
    if all_pts:
        all_pts = np.array(all_pts)
        pad = 0.6
        xmin, ymin = all_pts.min(axis=0) - pad
        xmax, ymax = all_pts.max(axis=0) + pad
        ax.set_xlim(xmin, xmax); ax.set_ylim(ymin, ymax)

    ax.set_aspect('equal', adjustable='box')
    ax.grid(True, ls='--', alpha=0.25)
    ax.set_title(f"最终时刻 t = {t:.0f} s 的布局")
    ax.set_xlabel("x / m"); ax.set_ylabel("y / m")
    ax.legend(loc='best')
    plt.tight_layout()

    if save_png:
        plt.savefig(png_path, dpi=300)
        print(f"[OK] 已保存：{png_path}")

    plt.show()

plot_final_layout()


