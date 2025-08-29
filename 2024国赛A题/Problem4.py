from __future__ import annotations
import math
from typing import Tuple, List
import numpy as np
from openpyxl import load_workbook

# ===== 基本参数（题目给定） =====
# 螺距（盘入=盘出）p = 1.7 m，等距螺线 r = b*θ
p = 1.7
b = p / (2.0*math.pi)

# 调头圆（问题3/4规定直径9m）
R_turn = 4.5

# 龙头速度（沿路径弧长）——已改为 1.12 m/s
V0 = 1.0  # m/s

# 板几何（孔心距）
L_head, L_body, D_hole = 3.41, 2.20, 0.55
l_HEAD = L_head - D_hole   # 2.86
l_BODY = L_body - D_hole   # 1.65
l_TAIL = l_BODY            # 1.65

# 节数：头1 + 身221 + 尾1
N_BODY = 221

# 时间网格（单位：秒）
T_START, T_END, DT = -100, 100, 1.0
T_GRID = np.arange(T_START, T_END+1e-12, DT, dtype=float)
# 速度差分步长（更小误差；单位：秒）
DT_V = 0.1

# ===== 基础向量工具 =====
def rot90(v: np.ndarray) -> np.ndarray:
    return np.array([-v[1], v[0]], float)

def angle_of(v: np.ndarray) -> float:
    return math.atan2(v[1], v[0])

def ang_diff(a_to: float, a_from: float, ccw: bool) -> float:
    d = (a_to - a_from) % (2.0*math.pi)
    return d if ccw else ((a_from - a_to) % (2.0*math.pi))

# ===== 螺线弧长原函数与反函数（只用于“时间↔θ”） =====
def F_theta(theta: float) -> float:
    # F(θ) = 0.5*( θ*sqrt(1+θ^2) + asinh(θ) )
    # 注意：等距螺线 r=bθ 的弧长 S(θ) = b * F(θ)
    return 0.5*(theta*math.sqrt(1.0+theta*theta) + math.asinh(theta))

def inv_F(Fv: float, max_iter=50, tol=1e-13) -> float:
    theta = math.sqrt(max(0.0, 2.0*Fv)) if Fv>1.0 else max(0.0, Fv)
    for _ in range(max_iter):
        f = F_theta(theta) - Fv
        df = math.sqrt(1.0 + theta*theta)
        theta_new = theta - f/df
        if theta_new < 0.0: theta_new = 0.0
        if abs(theta_new-theta) < tol: return theta_new
        theta = theta_new
    return theta

# ===== 盘入/盘出螺线 & 切向 =====
def xy_in(theta: float) -> np.ndarray:
    r = b*theta
    return np.array([r*math.cos(theta), r*math.sin(theta)], float)

def tan_in(theta: float) -> np.ndarray:
    # dP/dθ；盘入沿运动方向为顺时针（θ随时减小），取负
    dx = b*math.cos(theta) - b*theta*math.sin(theta)
    dy = b*math.sin(theta) + b*theta*math.cos(theta)
    v = np.array([-dx, -dy], float)
    n = np.hypot(*v);  return v/(n if n>1e-14 else 1.0)

def xy_out(theta: float) -> np.ndarray:
    r = b*(theta - math.pi)
    return np.array([r*math.cos(theta), r*math.sin(theta)], float)

def tan_out(theta: float) -> np.ndarray:
    u = theta - math.pi
    dx = b*math.cos(theta) - b*u*math.sin(theta)
    dy = b*math.sin(theta) + b*u*math.cos(theta)
    v = np.array([dx, dy], float)  # 盘出逆时针（θ随时增大）
    n = np.hypot(*v);  return v/(n if n>1e-14 else 1.0)

# ===== 触边点 B / F 及两圆几何（先大后小，R1:R2=2:1） =====
theta_B = R_turn / b
alpha = math.atan(theta_B)

# 两圆半径（来自你提供的推导：R1≈3/sinα, R2≈1.5/sinα）
R1 = 3.0 / math.sin(alpha)               # 大圆
R2 = 3.0 / (2.0*math.sin(alpha))         # 小圆

# B, F
B = xy_in(theta_B)
F = -B

# 圆心（你的截图公式）
C1 = np.array([ R_turn*math.cos(theta_B) - R1*math.sin(theta_B+alpha),
                R_turn*math.sin(theta_B) + R1*math.cos(theta_B+alpha) ], float)
C2 = np.array([-R_turn*math.cos(theta_B) + R2*math.sin(theta_B+alpha),
               -R_turn*math.sin(theta_B) - R2*math.cos(theta_B+alpha) ], float)

# 切向方向确定圆弧方向
tB = tan_in(theta_B)
tF = tan_out(theta_B + math.pi)

angB = angle_of(B - C1)
# CCW单位切向 = rot90(径向单位)
tan_ccw_at_B = rot90((B - C1)/np.hypot(*(B - C1)))
sgn1 = +1 if np.dot(tan_ccw_at_B, tB) > 0 else -1   # +1 表示沿CCW，-1 表示沿CW

# D 点：两圆外切线公切点（R1:R2=2:1 ⇒ D 在 C1→C2 方向上 2/3 处）
D = C1 + (R1/(R1+R2))*(C2 - C1)

angD_on_C1 = angle_of(D - C1)
phi1 = ang_diff(angD_on_C1, angB, ccw=(sgn1>0))
L1 = R1 * phi1  # 大弧弧长（米）

angD = angle_of(D - C2)
tan_ccw_at_F = rot90((F - C2)/np.hypot(*(F - C2)))
sgn2 = +1 if np.dot(tan_ccw_at_F, tF) > 0 else -1

angF_on_C2 = angle_of(F - C2)
phi2 = ang_diff(angF_on_C2, angD, ccw=(sgn2>0))
L2 = R2 * phi2  # 小弧弧长（米）

# ===== 段内坐标表达（圆弧参数 φ 从起点0开始量） =====
def xy_on_arc1(phi: float) -> np.ndarray:
    ang = angB + sgn1*phi
    return C1 + R1*np.array([math.cos(ang), math.sin(ang)], float)

def xy_on_arc2(phi: float) -> np.ndarray:
    ang = angD + sgn2*phi
    return C2 + R2*np.array([math.cos(ang), math.sin(ang)], float)

# ===== 同段“等距螺线”两点弦长为 l 的 Δθ 求解 =====
def same_spiral_delta(theta1: float, l: float, *, offset: float = 0.0) -> float:
    r1 = b*(theta1 - offset)
    if r1 < 1e-12: r1 = 1e-12
    def g(delta: float) -> float:
        r2 = b*(theta1 + delta - offset)
        return r1*r1 + r2*r2 - 2.0*r1*r2*math.cos(delta) - l*l
    def gp(delta: float) -> float:
        r2 = b*(theta1 + delta - offset)
        return 2.0*r2*b - 2.0*r1*b*math.cos(delta) + 2.0*r1*r2*math.sin(delta)
    # 牛顿+保守回缩
    delta = min(max(l/max(r1,1e-9), 1e-10), math.pi/2)
    ok = False
    for _ in range(20):
        val, der = g(delta), gp(delta)
        if abs(der) < 1e-14: break
        cand = delta - val/der
        if not (0.0 < cand < math.pi):
            cand = 0.5*(delta + max(1e-10, min(cand, math.pi-1e-10)))
        delta = cand
        if abs(val) < 1e-12: ok = True; break
    if ok: return delta
    # 二分兜底
    lo, hi = 1e-12, math.pi-1e-12
    gl, gh = g(lo), g(hi)
    if gl*gh > 0:
        hi = min(2*math.pi, hi + 1.0); gh = g(hi)
        if gl*gh > 0: return delta
    for _ in range(80):
        mid = 0.5*(lo+hi); gm = g(mid)
        if gm==0.0 or (hi-lo)<1e-12: return mid
        if gl*gm <= 0: hi, gh = mid, gm
        else:          lo, gl = mid, gm
    return 0.5*(lo+hi)

# ===== 头把手：t→(段, 段内参数)（关键：统一用 s=V0*t） =====
def head_state_at_time(t: float):
    """
    返回 (seg_id, param)：
      seg=1: 盘入螺线，param=θ_in
      seg=2: 大弧，     param=φ1 ∈ [0, L1/R1]
      seg=3: 小弧，     param=φ2 ∈ [0, L2/R2]
      seg=4: 盘出螺线，param=θ_out
    t=0 在 B；V0 为龙头沿路径的弧长速度（m/s）。
    """
    # 统一把时间 t（秒）映射为等效弧长 s（米）
    s = V0 * t

    if s <= 0.0:
        # 从 B 沿盘入往回：S(θ) = b*F(θ)
        # F(θ) = F(θ_B) + (-s)/b
        Fu = F_theta(theta_B) + (-s)/b
        theta = inv_F(Fu)
        return 1, theta

    if s <= L1:
        # 大弧：弧长 s = R1*φ1
        return 2, s / R1

    if s <= L1 + L2:
        # 小弧：弧长 s-L1 = R2*φ2
        return 3, (s - L1) / R2

    # 盘出：s_out = s - (L1 + L2)
    s_out = s - (L1 + L2)
    Fu = F_theta(theta_B) + s_out/b   # u = θ - π
    u = inv_F(Fu)
    theta = u + math.pi
    return 4, theta

# ===== 逐节回推（定弦长，自动跨段） =====
def step_prev(seg: int, param: float, l: float) -> Tuple[int, float]:
    if seg == 1:  # 盘入
        dth = same_spiral_delta(param, l, offset=0.0)
        return 1, param + dth

    elif seg == 2:  # 大弧
        # 同段直接Δφ
        dphi = 2.0*math.asin(min(1.0, l/(2.0*R1)))
        if param >= dphi + 1e-14:
            return 2, param - dphi
        # 跨到盘入：先吃到B
        chord_used = 2.0*R1*math.sin(max(0.0,param)/2.0)
        l_rem = max(0.0, l - chord_used)
        if l_rem <= 1e-14:
            return 1, theta_B
        dth = same_spiral_delta(theta_B, l_rem, offset=0.0)
        return 1, theta_B + dth

    elif seg == 3:  # 小弧
        dphi = 2.0*math.asin(min(1.0, l/(2.0*R2)))
        if param >= dphi + 1e-14:
            return 3, param - dphi
        # 先吃到D
        chord_used = 2.0*R2*math.sin(max(0.0,param)/2.0)
        l_rem = max(0.0, l - chord_used)
        if l_rem <= 1e-14:
            return 2, phi1
        # 进入大弧（从D往B）
        dphi1 = 2.0*math.asin(min(1.0, l_rem/(2.0*R1)))
        if phi1 >= dphi1 + 1e-14:
            return 2, phi1 - dphi1
        # 再跨入盘入
        chord_used2 = 2.0*R1*math.sin(max(0.0,phi1)/2.0)
        l_rem2 = max(0.0, l_rem - chord_used2)
        if l_rem2 <= 1e-14:
            return 1, theta_B
        dth = same_spiral_delta(theta_B, l_rem2, offset=0.0)
        return 1, theta_B + dth

    else:  # seg == 4, 盘出
        dth = same_spiral_delta(param, l, offset=math.pi)
        # 是否仍在盘出段
        if param - dth >= theta_B + math.pi - 1e-12:
            return 4, param - dth
        # 否则跨回 F
        P_now = xy_out(param)
        chord_to_F = float(np.hypot(*(P_now - F)))
        l_rem = max(0.0, l - chord_to_F)
        if l_rem <= 1e-14:
            return 3, phi2
        # 继续小弧（F→D 方向）
        dphi2 = 2.0*math.asin(min(1.0, l_rem/(2.0*R2)))
        if phi2 >= dphi2 + 1e-14:
            return 3, phi2 - dphi2
        # 极端再跨回大弧/盘入（通常用不到）
        chord_used2 = 2.0*R2*math.sin(max(0.0,phi2)/2.0)
        l_rem2 = max(0.0, l_rem - chord_used2)
        if l_rem2 <= 1e-14:
            return 2, phi1
        dphi1 = 2.0*math.asin(min(1.0, l_rem2/(2.0*R1)))
        if phi1 >= dphi1 + 1e-14:
            return 2, phi1 - dphi1
        chord_used3 = 2.0*R1*math.sin(max(0.0,phi1)/2.0)
        l_rem3 = max(0.0, l_rem - chord_used2 - chord_used3)
        if l_rem3 <= 1e-14:
            return 1, theta_B
        dth2 = same_spiral_delta(theta_B, l_rem3, offset=0.0)
        return 1, theta_B + dth2

# 给定（seg,param）取坐标
def xy_of(seg: int, param: float) -> np.ndarray:
    if seg==1:   return xy_in(param)
    if seg==2:   return xy_on_arc1(param)
    if seg==3:   return xy_on_arc2(param)
    return xy_out(param)

# ===== 生成某一时刻全队位置 =====
def all_handles_at_time(t: float) -> np.ndarray:
    # 头把手
    seg, par = head_state_at_time(t)
    P = [xy_of(seg, par)]
    # 第1节前（退 l_HEAD）
    seg, par = step_prev(seg, par, l_HEAD)
    P.append(xy_of(seg, par))
    # 第2..221节前
    for _ in range(2, N_BODY+1):
        seg, par = step_prev(seg, par, l_BODY)
        P.append(xy_of(seg, par))
    # 尾前
    seg, par = step_prev(seg, par, l_BODY)
    P.append(xy_of(seg, par))
    # 尾后
    seg, par = step_prev(seg, par, l_TAIL)
    P.append(xy_of(seg, par))
    return np.vstack(P)   # shape = (224, 2)

# ===== 速度（小步长中心差分；龙头速度强制为 V0） =====
def speeds_over_time(pos_series: np.ndarray, dt_s: float, dt_v: float = DT_V) -> np.ndarray:
    """
    pos_series: shape (T, N, 2), 按 1s 采样的坐标序列
    返回 speed: shape (T, N)
    """
    T, N, _ = pos_series.shape
    speed = np.zeros((T, N), float)

    # 用更密 dt_v 重新算邻点以求导（dt_v 单位：秒）
    def pos_at(t: float) -> np.ndarray:
        return all_handles_at_time(t)

    # 首末点：一侧差分；中间：中心差分
    for ti, t in enumerate(T_GRID):
        if t - T_START < 1e-12:           # 首点
            P2 = pos_at(t + dt_v)
            v = (P2 - pos_series[ti]) / dt_v
        elif T_END - t < 1e-12:           # 末点
            P0 = pos_at(t - dt_v)
            v = (pos_series[ti] - P0) / dt_v
        else:                              # 中心差分
            P2 = pos_at(t + dt_v)
            P0 = pos_at(t - dt_v)
            v = (P2 - P0) / (2.0*dt_v)
        speed[ti, :] = np.hypot(v[:,0], v[:,1])
        # 龙头速度置为 V0（消除数值误差）
        speed[ti, 0] = V0
    return speed

# ===== 主过程：计算并写入 result4.xlsx =====
def make_result4(xlsx_path: str):
    # 1) 位置序列（每秒一次）
    all_pos = []
    print("[INFO] Generating positions ...")
    for t in T_GRID:
        all_pos.append(all_handles_at_time(t))
    pos_arr = np.stack(all_pos, axis=0)       # (T, 224, 2)

    # 2) 速度序列（小步长差分）
    print("[INFO] Differentiating for speeds ...")
    spd_arr = speeds_over_time(pos_arr, dt_s=DT, dt_v=DT_V)  # (T, 224)

    # 3) 打开模板并写入
    print(f"[INFO] Writing to: {xlsx_path}")
    wb = load_workbook(xlsx_path)
    ws_pos = wb["位置"]
    ws_spd = wb["速度"]

    Tn, Npts = pos_arr.shape[0], pos_arr.shape[1]
    # 逐时刻列写入（与您之前模板一致：第2列开始是各时刻；行2起每两行一组 x/y）
    for ti in range(Tn):
        col = 2 + ti
        for j in range(Npts):
            base = 2 + 2*j
            ws_pos.cell(row=base,   column=col, value=round(float(pos_arr[ti, j, 0]), 6))
            ws_pos.cell(row=base+1, column=col, value=round(float(pos_arr[ti, j, 1]), 6))
            ws_spd.cell(row=1+1+j,  column=col, value=round(float(spd_arr[ti, j]), 6))

    wb.save(xlsx_path)
    print("[DONE] result4.xlsx written.")

    # 4) 打印若干几何与一致性检查
    print("\n[Geom]")
    print(f"  V0={V0:.3f} m/s")  # 新增：打印龙头速度
    print(f"  theta_B={theta_B:.9f}, alpha={alpha:.9f}")
    print(f"  C1=({C1[0]:.6f},{C1[1]:.6f}), R1={R1:.6f}, phi1={phi1:.9f}, L1={L1:.6f}")
    print(f"  C2=({C2[0]:.6f},{C2[1]:.6f}), R2={R2:.6f}, phi2={phi2:.9f}, L2={L2:.6f}")
    print(f"  B=({B[0]:.6f},{B[1]:.6f}), D=({D[0]:.6f},{D[1]:.6f}), F=({F[0]:.6f},{F[1]:.6f})")

    # 距离约束检查（随机抽几帧）
    for tchk in (-100, 0, 100):
        k = int((tchk - T_START)/DT + 0.5)
        P = pos_arr[k]
        errs = []
        # 头->1节
        errs.append(abs(np.hypot(*(P[0]-P[1])) - l_HEAD))
        # 身体相邻（1..221 与下一节）
        for i in range(1, 1+N_BODY-1):
            errs.append(abs(np.hypot(*(P[i]-P[i+1])) - l_BODY))
        # 最后一节身前 -> 尾前
        errs.append(abs(np.hypot(*(P[1+N_BODY-1]-P[1+N_BODY])) - l_BODY))
        # 尾前 -> 尾后
        errs.append(abs(np.hypot(*(P[1+N_BODY]-P[1+N_BODY+1])) - l_TAIL))
        print(f"[CHECK] t={tchk:>4}s  max|ΔL|={max(errs):.6e} m, mean|ΔL|={np.mean(errs):.6e} m")

if __name__ == "__main__":
    # 改成你的模板路径
    xlsx_path = r"D:/MathModeling/2024国赛A题/附件/result4.xlsx"
    make_result4(xlsx_path)




